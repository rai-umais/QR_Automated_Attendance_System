import openpyxl
import os
from datetime import datetime
import re

def normalize_roll(roll):
    if not roll: return ""
    # Extract digits to compare (e.g., 22L-7998 -> 227998, l227998 -> 227998)
    return "".join(re.findall(r'\d+', str(roll)))

def export_attendance_to_xlsx(date, present_roll_numbers, xlsx_path, enrolled_students):
    """
    Updates an existing attendance sheet with a new column for the given date.
    Marks 'P' for present and 'A' for absent.
    """
    # If the master file doesn't exist, we fallback to creating a new one
    if not os.path.exists(xlsx_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.cell(1, 1).value = "Roll Number"
        ws.cell(1, 2).value = "Student Name"
        ws.cell(1, 3).value = "Email"
        start_row = 2
        
        # Populate students if they are missing
        for idx, student in enumerate(enrolled_students, start=2):
            ws.cell(idx, 1).value = student.roll_number
            ws.cell(idx, 2).value = student.name
            ws.cell(idx, 3).value = student.email
        next_col = 4
    else:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        next_col = ws.max_column + 1
        start_row = 2

    # Set Date Header (e.g., F1, G1...)
    date_str = date.strftime('%d/%m/%Y')
    ws.cell(1, next_col).value = date_str
    
    # Normalize present rolls for comparison
    present_norms = {normalize_roll(r) for r in present_roll_numbers}

    # Update the column with P/A
    # We iterate through the rows in the sheet to find roll numbers
    # We'll check column 2 (Roll No. in user's file)
    for row in range(start_row, ws.max_row + 1):
        # Check column 2 (B) for Roll Number
        roll_val = ws.cell(row, 2).value
        if not roll_val:
            # If column 2 is empty, check column 1 (just in case)
            roll_val = ws.cell(row, 1).value
        
        if roll_val:
            norm_roll = normalize_roll(roll_val)
            if norm_roll:
                ws.cell(row, next_col).value = 'P' if norm_roll in present_norms else 'A'

    # Save the updated workbook
    wb.save(xlsx_path)
    return True
