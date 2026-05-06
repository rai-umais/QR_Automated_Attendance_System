from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Iterable
import re
import zipfile

import openpyxl


@dataclass
class ParsedStudent:
    roll_number: str
    name: str
    email: str


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = normalized.replace("_", " ")
    normalized = re.sub(r"[^a-z0-9\s]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _pick_column(headers: dict[str, int], aliases: Iterable[str]) -> int | None:
    for alias in aliases:
        idx = headers.get(alias)
        if idx is not None:
            return idx
    return None


def _normalize_roll_number(value: object) -> str:
    if value is None:
        return ""
    roll = str(value).strip().lower()
    # Remove dashes to keep roll format and generated email consistent.
    return roll.replace("-", "")


def parse_students_from_xlsx(file_bytes: bytes) -> list[ParsedStudent]:
    """
    Parse an Excel file and return student rows with roll_number + name.
    If email is missing, generate: <roll_number>@lhr.nu.edu.pk.
    """
    if not file_bytes:
        raise ValueError("Uploaded file is empty.")
    if not zipfile.is_zipfile(BytesIO(file_bytes)):
        raise ValueError("Invalid Excel workbook. Please upload a real .xlsx file.")

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    header_row_number = None
    headers: dict[str, int] = {}
    roll_idx = None
    name_idx = None
    email_idx = None

    for row_no, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(15, sheet.max_row), values_only=True),
        start=1
    ):
        row_headers = {_normalize_header(cell): idx for idx, cell in enumerate(row)}
        maybe_roll = _pick_column(
            row_headers,
            (
                "roll number",
                "roll no",
                "roll",
                "rollnumber",
                "registration number",
                "reg no",
            ),
        )
        maybe_name = _pick_column(
            row_headers,
            ("student name", "name", "student"),
        )
        if maybe_roll is not None and maybe_name is not None:
            header_row_number = row_no
            headers = row_headers
            roll_idx = maybe_roll
            name_idx = maybe_name
            email_idx = _pick_column(headers, ("email", "email address", "mail"))
            break

    if header_row_number is None or roll_idx is None or name_idx is None:
        raise ValueError(
            "Excel must contain roll and name headers (e.g., 'Roll No.' and 'Student Name')."
        )

    parsed: list[ParsedStudent] = []
    seen_pairs: set[tuple[str, str]] = set()
    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        roll_raw = row[roll_idx] if roll_idx < len(row) else None
        name_raw = row[name_idx] if name_idx < len(row) else None
        email_raw = row[email_idx] if email_idx is not None and email_idx < len(row) else None

        roll = _normalize_roll_number(roll_raw)
        name = str(name_raw).strip() if name_raw is not None else ""

        if not roll or not name:
            continue

        key = (roll, name.casefold())
        if key in seen_pairs:
            continue
        seen_pairs.add(key)

        email = str(email_raw).strip().lower() if email_raw is not None else ""
        if not email:
            email = f"{roll}@lhr.nu.edu.pk"

        parsed.append(ParsedStudent(roll_number=roll, name=name, email=email))

    return parsed
